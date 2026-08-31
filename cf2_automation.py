import os
import re
import sys
from datetime import datetime
import openpyxl
from cf2_mapper import build_cf2_data
import cf2_api
from cf2_fees import get_fees
from draft_automation import (
    run_create_draft_flow,
    InvalidMemberPinError,
)
from draft_title import build_draft_title

# Guard against print() crashing an otherwise-successful patient run.
# When this runs under a packaged .exe, stdout/stderr can be attached to a
# legacy Windows console codepage (cp1252, cp437, etc.) instead of UTF-8.
# Any print() containing a character outside that codepage (e.g. "✓", "—")
# raises UnicodeEncodeError ('charmap' codec can't encode character ...),
# which — since it's unhandled at the point of the print() call itself —
# aborts the whole patient rather than just producing a garbled log line.
# Setting errors="replace" keeps each stream's existing encoding but swaps
# unencodable characters for a placeholder instead of raising.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(errors="replace")
    except (AttributeError, ValueError):
        pass




def resource_path(relative_path):
    """
    Returns the correct path both for development
    and for the packaged PyInstaller executable.
    """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


CF2_TEMPLATE_PATH = resource_path(
    os.path.join(
        "templates",
        "CF2_Template.xlsx"
    )
)


class TransmittalNotFoundError(Exception):
    """
    Raised when mode="existing_draft" and the transmittal number from the
    uploaded workbook doesn't match anything in Beacon's Transmittals
    list.

    Distinct from a bare Exception for the same reason InvalidMemberPinError
    is (see its docstring): process_patient() needs to tell "this row's
    transmittal number is wrong/doesn't exist yet, skip it and move on"
    apart from an unexpected automation/API error that should mark the row
    failed.
    """
    pass


class CF2Automation:

    def __init__(self, uploaded_excel_path=None, mode="new_draft"):
        # Path to the workbook the user actually uploaded. Sheet2 (Billing
        # Clerk / Accountant name, contact no., official capacity) must be
        # read from THIS file, not from the bundled CF2_TEMPLATE_PATH.
        self.uploaded_excel_path = uploaded_excel_path or CF2_TEMPLATE_PATH
        # "new_draft" (default): fill_cf2() creates a brand-new Beacon
        # draft via Create Draft + Add Claims, using each row's Member
        # PIN — the original behavior, unchanged.
        # "existing_draft": fill_cf2() skips draft creation entirely and
        # instead searches the Transmittals list for each row's
        # transmittal number (already created in Beacon some other way),
        # then continues into the exact same CF2 field-filling steps.
        self.mode = mode
        # Every processed patient gets one entry here:
        # {"transmittal": ..., "patient_name": ..., "status": "success"/"skipped"/"failed", "message": ...}
        self.results = []
        # API-resolved IDs for the patient currently being processed.
        # Both new-draft and existing-draft modes set these directly from API
        # responses; no browser/page URL is involved.
        self._current_ids = None

    # ------------------------------------------------------------------
    # Public entry point — never lets a single patient crash the batch
    # ------------------------------------------------------------------
    def process_patient(self, record):
        result = {
            "transmittal": getattr(record, "transmittal", "?"),
            "patient_name": getattr(record, "patient_name", "?"),
            "status": "failed",
            "message": "",
        }

        try:
            data = build_cf2_data(record)
        except Exception as e:
            result["message"] = f"Could not build CF2 data: {e}"
            print(f"ERROR: {result['message']}")
            self.results.append(result)
            return result

        result["transmittal"] = data.transmittal
        result["patient_name"] = data.patient_name

        print("=" * 50)
        print(f"PROCESSING PATIENT: {data.patient_name}  (Transmittal: {data.transmittal})")
        print("=" * 50)

        try:
            status, message = self.fill_cf2(data)
            result["status"] = status
            result["message"] = message
            # fill_cf2 overwrites data.transmittal with the auto-generated
            # number once the draft is created, so re-read it here.
            result["transmittal"] = data.transmittal
        except (InvalidMemberPinError, TransmittalNotFoundError) as e:
            # Either the PIN was bad (new_draft mode) or the transmittal
            # number didn't match anything (existing_draft mode) — not an
            # automation glitch either way. Safe to record this row as
            # skipped and move on to the next one rather than treating it
            # as a hard failure.
            result["status"] = "skipped"
            result["message"] = str(e)
            print(f"SKIPPED: {data.patient_name} — {e} — moving to next row.")
        except Exception as e:
            result["status"] = "failed"
            result["message"] = f"Unhandled error: {e}"
            print(f"ERROR: Unhandled exception for {data.patient_name}: {e}")

        self.results.append(result)
        return result

    def get_summary(self):
        """Returns the list of per-patient results collected so far."""
        return self.results

    # ------------------------------------------------------------------
    # Teardown — call this once after all patients are processed
    # ------------------------------------------------------------------
    def close(self):
        """Compatibility no-op.

        CF2 automation is API-only and owns no browser/page resources.
        Existing callers may continue invoking close().
        """
        return None

    # ------------------------------------------------------------------
    # Excel lookup — "Prepared by" (Billing Clerk / Accountant) name
    # ------------------------------------------------------------------
    def _get_billing_clerk_name(self):
        """Reads the Billing Clerk / Accountant name from Sheet2!A1 of the
        UPLOADED workbook (falls back to CF2_TEMPLATE_PATH only if no file
        was uploaded)."""
        wb = openpyxl.load_workbook(self.uploaded_excel_path, data_only=True)
        sheet2 = wb["Sheet2"]
        value = sheet2["A1"].value
        wb.close()
        return str(value).strip() if value is not None else ""

    def _get_billing_clerk_cp(self):
        wb = openpyxl.load_workbook(self.uploaded_excel_path, data_only=True)
        sheet2 = wb["Sheet2"]
        value = sheet2["A2"].value
        wb.close()
        return str(value).strip() if value else ""

    def _get_official_capacity_designation(self):
        """Reads the Official Capacity/Designation from Sheet2!B1 of the
        UPLOADED workbook (falls back to CF2_TEMPLATE_PATH only if no file
        was uploaded)."""
        wb = openpyxl.load_workbook(self.uploaded_excel_path, data_only=True)
        sheet2 = wb["Sheet2"]
        value = sheet2["B1"].value
        wb.close()
        return str(value).strip() if value is not None else ""

    def _get_ids(self):
        """Return API-resolved IDs for the patient currently being processed."""
        if self._current_ids is None:
            raise RuntimeError(
                "Current transmittal/claim IDs were not resolved through the API."
            )
        return self._current_ids

    # ------------------------------------------------------------------
    # Step runner
    # ------------------------------------------------------------------
    def _step(self, description, func, critical=True):
        """
        Runs a single automation step.
        - critical=True : failure aborts the current patient (raises up to
          process_patient, which marks the patient as failed and moves on
          to the next one).
        - critical=False: failure is logged as a warning and swallowed so
          the rest of the CF2 form can still be attempted.
        """
        try:
            print(description)
            func()
            return True
        except Exception as e:
            print(f"WARNING: Step failed [{description}]: {e}")
            if critical:
                raise
            return False

    # ------------------------------------------------------------------
    # Main flow
    # ------------------------------------------------------------------
    def fill_cf2(self, data):
        self._current_ids = None

        if self.mode == "existing_draft":
            self._locate_existing_draft(data)
        else:
            self._create_draft(data)

        self._step(
            "Validating eligibility via API...",
            lambda: self._validate_eligibility_via_api(data),
            critical=False,
        )

        self._add_discharge_diagnosis()
        self._add_surgical_procedure(data)
        self._add_doctor(data)

        if not self._fill_and_save_cf2(data, persist_session_dates=True):
            raise RuntimeError(
                "Could not persist CF2/session dates through EditPHICCF2 API."
            )

        ids = self._get_ids()
        self._step(
            "Tagging Surgical Procedure as 1st Case Rate via API...",
            lambda: self._tag_first_case_via_api(ids, data),
            critical=True,
        )

        # Keep the second save from the proven flow. It writes the final CF2
        # values after case-rate tagging without changing the established logic.
        if not self._fill_and_save_cf2(data):
            raise RuntimeError("Could not save CF2 through EditPHICCF2 API.")

        self._diagnose_session_date_backend(data, "AFTER CF2 SAVE")
        self._fill_statement_of_account(data)

        print("SUCCESS: CF2 completed for this patient.")
        return "success", "CF2 completed successfully."

    # ------------------------------------------------------------------
    # Steps
    # ------------------------------------------------------------------
    def _create_draft(self, data):
        """
        Creates the Beacon transmittal (Create Draft) and runs Add Claims
        for this patient's Member PIN, admission date (= first treatment
        date) and discharge date (= last treatment date), with an
        auto-generated draft title. The API response supplies the generated
        transmittal/claim IDs directly, so fill_cf2() continues without any
        browser navigation or URL parsing.
        """
        admission_date = data.first_treatment.strftime("%m/%d/%Y")
        discharge_date = data.last_treatment.strftime("%m/%d/%Y")
        draft_title = build_draft_title(data.patient_name, data.first_treatment, data.last_treatment)

        result = run_create_draft_flow(
            None, data.member_pin, admission_date, discharge_date, draft_title
        )
        self._current_ids = {
            "transmittal_id": int(result["transmittal_id"]),
            "claim_id": int(result["claim_id"]),
        }
        data.transmittal = result.get("transmittal_number") or "AUTO-GENERATED"
        print(f"Draft created. Transmittal number: {data.transmittal}")

    # ------------------------------------------------------------------
    # Existing-draft flow — used when self.mode == "existing_draft".
    # Skips Create Draft + Add Claims entirely and instead locates a
    # transmittal that already exists in Beacon by number, then hands
    # off to the exact same CF2 field-filling steps fill_cf2() runs for
    # a freshly-created draft. One transmittal maps to exactly one
    # patient, so grabbing the first (only) row at each step is correct.
    # ------------------------------------------------------------------
    def _locate_existing_draft(self, data):
        """Resolve an existing draft through Beacon APIs only.

        Selection logic is intentionally unchanged: search using
        ``data.transmittal`` and process the first claim in that transmittal.
        """
        def _run():
            resolved = cf2_api.resolve_existing_draft(data.transmittal, attempts=3)
            if resolved is None:
                raise TransmittalNotFoundError(
                    f"Transmittal not found: {data.transmittal}"
                )
            self._current_ids = {
                "transmittal_id": resolved["transmittal_id"],
                "claim_id": resolved["claim_id"],
            }

        self._step(
            f"Locating existing draft via API (Transmittal: {data.transmittal})...",
            _run,
            critical=True,
        )
        print(
            f"Located existing draft via API. Transmittal number: {data.transmittal}; "
            f"transmittalId={self._current_ids['transmittal_id']}, "
            f"claimId={self._current_ids['claim_id']}"
        )




    def _validate_eligibility_via_api(self, data):
        """Run Beacon's Validate Eligibility workflow through APIs only."""
        ids = self._get_ids()
        result = cf2_api.validate_claim_eligibility(
            claim_id=ids["claim_id"],
            transmittal_id=ids["transmittal_id"],
            admission_date=data.first_treatment,
            discharge_date=data.last_treatment,
        )
        if result.get("skipped"):
            print("Eligibility already validated in CF1 - skipping duplicate PBEF generation.")
            return True

        saved = result.get("saved") or {}
        print(
            "Eligibility validated via API: "
            f"status='{saved.get('claimStatusDescription')}', "
            f"eligibleAsOf='{saved.get('eligibleAsOf')}', "
            f"remainingDays='{saved.get('eligibilityRemainingDays')}'."
        )
        return True


    def _fill_and_save_cf2(self, data, persist_session_dates=False):
        """
        Saves the CF2 record through the confirmed API path
        (_fill_referral_and_accommodation, _fill_disposition_and_diagnosis,
        _fill_benefits_and_fees, _fill_access_patient_records_date,
        _save_cf2) into a single GET -> merge -> PUT API call, confirmed
        via a full (uncut) HAR capture of a real CF2 save.

        Beacon expects the WHOLE CF2 record on every save, not a partial
        patch — fields this migration doesn't know about (newborn care,
        TB DOTS, animal bite, cataract, surgicalProcedures, audit
        fields, etc.) live in this same record. So this always starts
        from cf2_api.get_cf2()'s current state and only mutates the
        specific CF2 fields this automation owns, leaving
        everything else exactly as the server returned it — the only
        way to do this safely without risking silently wiping data
        outside this migration's scope.

        Returns True on success and False if the API save fails.
        """
        try:
            ids = self._get_ids()
            cf2_record = cf2_api.get_cf2(ids["claim_id"])
            if not cf2_record:
                raise cf2_api.Cf2ApiError(
                    f"GetPHICCF2ById returned nothing for "
                    f"claimId={ids['claim_id']}."
                )

            # _fill_referral_and_accommodation
            cf2_record["isPatientReferred"] = "N"
            cf2_record["accomodationTypeCode"] = "P"
            cf2_record["accomodationTypeValue"] = "Private"

            # Preserve Beacon's existing patient type; only supply the required
            # outpatient default when the API record does not contain one.
            if not cf2_record.get("patientTypeCode"):
                cf2_record["patientTypeCode"] = "O"
                cf2_record["patientTypeValue"] = "Outpatient"

            # Admission stays at local midnight; discharge moves to
            # local noon (AM -> PM) — confirmed via HAR, see
            # to_utc_midnight_iso / to_utc_noon_iso docstrings.
            admission_iso = cf2_api.to_utc_midnight_iso(data.first_treatment)
            discharge_iso = cf2_api.to_utc_noon_iso(data.last_treatment)
            cf2_record["admissionDate"] = data.first_treatment.strftime("%m-%d-%Y")
            cf2_record["admissionDateTime"] = admission_iso
            cf2_record["admissionTime"] = admission_iso
            cf2_record["dischargeDate"] = data.last_treatment.strftime("%m-%d-%Y")
            cf2_record["dischargeDateTime"] = discharge_iso
            cf2_record["dischargeTime"] = discharge_iso

            # _fill_disposition_and_diagnosis
            # preserve the established CF2 behavior.
            cf2_record["patientDispositionCode"] = "I"
            cf2_record["patientDispositionValue"] = "Improved"
            cf2_record["admissionDiagnosis"] = "CHRONIC KIDNEY DISEASE STAGE V"

            # _fill_benefits_and_fees
            fees = get_fees(data.total_sessions)
            cf2_record["doesPatientHasEnoughBenefits"] = "N"
            cf2_record["hospitalFeesActualCharges"] = f"{fees['hospital_actual']:.2f}"
            cf2_record["hospitalFeesAmountAfterDiscount"] = f"{fees['hospital_discount']:.2f}"
            cf2_record["hospitalFeesPhilHealthBenefit"] = f"{fees['hospital_discount']:.2f}"
            cf2_record["professionalFeesActualCharges"] = f"{fees['prof_actual']:.2f}"
            cf2_record["professionalFeesAmountAfterDiscount"] = f"{fees['prof_discount']:.2f}"
            cf2_record["professionalFeesPhilHealthBenefit"] = f"{fees['prof_discount']:.2f}"
            cf2_record["hospitalFeesDidPatientPay"] = "N"
            cf2_record["hospitalFeesPatientHasHMO"] = "N"
            cf2_record["hospitalFeesPatientHasOtherDeductions"] = "N"
            cf2_record["professionalFeesDidPatientPay"] = "N"
            cf2_record["professionalFeesPatientHasHMO"] = "N"
            cf2_record["professionalFeesPatientHasOtherDeductions"] = "N"
            cf2_record["purchasesWithDrugsMedSupplies"] = "N"
            cf2_record["purchasesWithExaminations"] = "N"

            print(
                f"Fees for {data.total_sessions} sessions: "
                f"hosp={fees['hospital_actual']}/{fees['hospital_discount']}, "
                f"prof={fees['prof_actual']}/{fees['prof_discount']}"
            )

            # _fill_access_patient_records_date
            cf2_record["aprDate"] = data.last_treatment.strftime("%m-%d-%Y")

            # GetPHICCF2ById does NOT embed surgicalProcedures (confirmed
            # via live debugging - see cf2_api.get_cf2()'s docstring),
            # but EditPHICCF2 requires it to be present or it 500s
            # unconditionally, regardless of every other field. Fetch it
            # separately and reshape it to match a real captured working
            # save payload before sending.
            cf2_record["surgicalProcedures"] = (
                cf2_api.build_surgical_procedures_for_cf2(ids["claim_id"])
            )

            # When requested, persist session dates through the full EditPHICCF2
            # payload. This mirrors Beacon's successful manual request: the
            # session object keeps its id/session metadata and only sessionDate
            # is changed. Do NOT try to persist this through NewPHICAllCaseRate;
            # that endpoint returns a case-rate record and does not return or
            # reliably store sessions[].sessionDate.
            if persist_session_dates and data.session_dates:
                procedures = cf2_record.get("surgicalProcedures") or []
                if not procedures:
                    raise RuntimeError("EditPHICCF2 payload contains no surgicalProcedures.")

                if len(procedures) < 1:
                    raise RuntimeError("No Surgical Procedure available for session-date persistence.")

                for proc_index, procedure in enumerate(procedures):
                    sessions = procedure.get("sessions") or []
                    if not sessions:
                        raise RuntimeError(
                            f"Surgical Procedure #{proc_index + 1} has no session objects "
                            "in the EditPHICCF2 payload; refusing to invent session IDs."
                        )
                    for session_index, session in enumerate(sessions):
                        if session_index >= len(data.session_dates):
                            break
                        session["sessionDate"] = data.session_dates[session_index].strftime("%m-%d-%Y")
                        print(
                            f"  API Session {session_index + 1}: "
                            f"sessionDate={session['sessionDate']} "
                            f"(session id={session.get('id')})"
                        )

                print("Saving session date(s) through EditPHICCF2 API...")

            cf2_api.edit_cf2(cf2_record)
            print("CF2 form fields filled and saved (via API).")

            if persist_session_dates and data.session_dates:
                # Verify the value returned by Beacon, not merely the request
                # body. A successful HTTP response alone is insufficient.
                verify_procedures = cf2_api.get_surgical_procedures(ids["claim_id"])
                expected_dates = [d.strftime("%m-%d-%Y") for d in data.session_dates]
                actual_dates = []
                for procedure in verify_procedures or []:
                    for session in (procedure.get("sessions") or []):
                        actual_dates.append(session.get("sessionDate"))

                actual_dates = actual_dates[:len(expected_dates)]
                print("===== SESSION DATE API VERIFICATION (EditPHICCF2) =====")
                for idx, expected in enumerate(expected_dates):
                    actual = actual_dates[idx] if idx < len(actual_dates) else None
                    print(
                        f"  Session {idx + 1}: expected={expected!r}, "
                        f"backend={actual!r}"
                    )
                if [self._normalize_backend_session_date(v) for v in actual_dates] != expected_dates:
                    raise RuntimeError(
                        f"EditPHICCF2 returned session dates {actual_dates!r}; "
                        f"expected {expected_dates!r}."
                    )
                print("RESULT: SESSION DATE(S) PERSISTED AND VERIFIED BY BACKEND API.")
                print("========================================================")

            return True
        except Exception as e:
            print(
                f"WARNING: API path for CF2 field save failed ({e}) - "
                f"API-only CF2 save aborted."
            )
            return False




    def _add_discharge_diagnosis(self):
        ids = self._get_ids()
        existing = cf2_api.get_discharge_diagnoses(ids["claim_id"])
        if existing:
            print("Discharge Diagnosis already exists (via API) - skipping.")
            return
        cf2_api.add_discharge_diagnosis_n18_5(ids["claim_id"])
        print("Discharge Diagnosis added and set as Primary (via API).")

        








    def _add_surgical_procedure(self, data):
        ids = self._get_ids()
        existing = cf2_api.get_surgical_procedures(ids["claim_id"])
        if existing:
            print("Surgical Procedure already exists (via API) - skipping creation.")
            return

        icd10_matches = cf2_api.search_icd10("N18.5")
        if not icd10_matches:
            raise cf2_api.Cf2ApiError("SearchICD10('N18.5') returned no matches.")

        cf2_api.new_surgical_procedure(
            ids["claim_id"],
            icd10_matches[0]["icD10Code"],
            icd10_matches[0]["icD10Value"],
            data.total_sessions,
        )
        print("Surgical Procedure created (via API).")

    @staticmethod
    def _normalize_backend_session_date(value):
        """Normalize Beacon sessionDate values for date-only comparison.

        Beacon may return the same persisted date as either MM-DD-YYYY or
        an ISO datetime such as 2026-07-01T00:00:00. Compare calendar dates,
        not their raw string representations.
        """
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
            try:
                return datetime.strptime(text, fmt).strftime("%m-%d-%Y")
            except ValueError:
                continue
        # Handle ISO timestamps with timezone/offset by using only the date
        # portion, which is how Beacon represents this field in its API.
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).strftime("%m-%d-%Y")
        except ValueError:
            return text[:10]

    def _diagnose_session_date_backend(self, data, stage):
        """Best-effort diagnostic: print Beacon's persisted sessionDate.

        This deliberately does not modify anything. It distinguishes a
        a transient client value from a value actually returned by
        GetPHICSurgicalProcedure, and helps identify whether a later
        operation wipes the session date.
        """
        try:
            ids = self._get_ids()
            procedures = cf2_api.get_surgical_procedures(ids["claim_id"])
            print(f"===== SESSION DATE BACKEND DIAGNOSTIC: {stage} =====")
            if not procedures:
                print("  Surgical Procedure records returned: 0")
                print("  Session 1 backend sessionDate: <NO PROCEDURE>")
                print("=======================================================")
                return

            for proc_index, procedure in enumerate(procedures, start=1):
                sessions = procedure.get("sessions") or []
                print(
                    f"  Surgical Procedure #{proc_index}: "
                    f"id={procedure.get('id')}, "
                    f"numberOfSessions={procedure.get('numberOfSessions', len(sessions))}"
                )
                if not sessions:
                    print("    sessions: []")
                    continue
                for session_index, session in enumerate(sessions, start=1):
                    print(
                        f"    Session {session_index} backend sessionDate: "
                        f"{session.get('sessionDate')!r}"
                    )

            expected = (
                data.session_dates[0].strftime("%m-%d-%Y")
                if data.session_dates
                else "<NONE>"
            )
            actual = None
            first_sessions = procedures[0].get("sessions") or []
            if first_sessions:
                actual = first_sessions[0].get("sessionDate")

            print(f"  Expected session 1 date from CF2 data: {expected!r}")
            print(f"  Actual session 1 backend date: {actual!r}")
            if self._normalize_backend_session_date(actual) == expected:
                print("  DIAGNOSTIC RESULT: BACKEND DATE MATCHES EXPECTED DATE.")
            elif actual:
                print(
                    "  DIAGNOSTIC RESULT: BACKEND HAS A DATE, BUT IT DOES NOT "
                    "MATCH EXPECTED."
                )
            else:
                print("  DIAGNOSTIC RESULT: BACKEND SESSION DATE IS EMPTY/NULL.")
            print("=======================================================")
        except Exception as e:
            print(
                f"WARNING: Session date backend diagnostic failed at {stage}: {e}"
            )

    def _tag_first_case_via_api(self, ids, data):
        """Tag the hemodialysis Surgical Procedure as 1st Case Rate via API.

        Multi-session behavior is confirmed by HAR: one NewPHICAllCaseRate
        request contains every session and its MM-DD-YYYY date.
        """
        try:
            if cf2_api.get_case_rates(ids["claim_id"]):
                print("Surgical Procedure already tagged as 1st Case Rate (via API) - skipping.")
                return

            procedures = cf2_api.get_surgical_procedures(ids["claim_id"])
            if not procedures:
                print("No Surgical Procedure found - can't tag via API.")
                raise RuntimeError("No Surgical Procedure found for 1st Case Rate tagging.")
            surgical_procedure = procedures[0]
            sessions = surgical_procedure.get("sessions") or []
            if len(sessions) < len(data.session_dates):
                raise RuntimeError(
                    f"Backend has {len(sessions)} session row(s), expected at least "
                    f"{len(data.session_dates)}."
                )

            expected_dates = [d.strftime("%m-%d-%Y") for d in data.session_dates]
            actual_before = [s.get("sessionDate") for s in sessions[:len(expected_dates)]]
            normalized_before = [self._normalize_backend_session_date(v) for v in actual_before]
            if normalized_before != expected_dates:
                raise RuntimeError(
                    f"Refusing 1st Case Rate tagging: backend session dates are "
                    f"{actual_before!r}, expected {expected_dates!r}."
                )

            hospital_identity = cf2_api.get_hospital_identity(ids["transmittal_id"])
            target_date_str = data.session_dates[0].strftime("%m-%d-%Y")
            case_rate_response = cf2_api.search_case_rates(
                "90935", target_date_str, hospital_identity
            )
            caserates = case_rate_response.get("caserates") or []
            if not caserates:
                raise cf2_api.Cf2ApiError(
                    f"SearchCaseRates returned no case rates for 90935 / {target_date_str}."
                )

            self._diagnose_session_date_backend(data, "BEFORE 1ST CASE TAG")
            cf2_api.tag_first_case(
                ids["claim_id"], surgical_procedure, caserates[0], data.session_dates
            )
            print(
                f"Surgical Procedure tagged as 1st Case Rate via API "
                f"with {len(data.session_dates)} session(s)."
            )

            procedures_after = cf2_api.get_surgical_procedures(ids["claim_id"])
            sessions_after = (procedures_after[0].get("sessions") or []) if procedures_after else []
            actual_after = [
                s.get("sessionDate") for s in sessions_after[:len(expected_dates)]
            ]
            normalized_after = [self._normalize_backend_session_date(v) for v in actual_after]
            self._diagnose_session_date_backend(data, "AFTER 1ST CASE TAG")
            if normalized_after != expected_dates:
                raise RuntimeError(
                    f"1st Case Rate tagging changed/lost session dates: "
                    f"expected {expected_dates!r}, got {actual_after!r}."
                )

            return
        except Exception as e:
            raise RuntimeError(f"API 1st Case Rate tagging failed: {e}") from e

    def _fill_claim_form_two_via_api(self, claim_id, data):
        """
        Save the Claim Form 2 PDF signature block entirely through Beacon's
        API without opening the Claim Form 2 page.

        Beacon's GET /api/PHICDocument/GetCf2PdfDetails endpoint is used only
        to obtain the complete current PDF data object. The Part III patient/member
        name fields and Part IV HCI representative fields are then changed, and
        the complete object is submitted
        to POST /api/PHICDocument/NewPdfClaimFormTwo.

        IMPORTANT: GetCf2PdfDetails is NOT used as a post-save verification
        source. A real successful save returned HTTP 200 from
        NewPdfClaimFormTwo, while a subsequent GetCf2PdfDetails continued to
        return some rendered PDF fields as null. The authoritative
        verification for this endpoint is the generated Claim Form 2 PDF,
        which was observed containing the submitted signature/designation/date.

        NewPdfClaimFormTwo returns an empty JSON response body on success, so
        _post() treating the successful response as None is expected.
        """
        client_id = cf2_api.get_client_id()
        base = cf2_api.get_cf2_pdf_details(claim_id, client_id)
        if not isinstance(base, dict) or not base:
            raise cf2_api.Cf2ApiError(
                f"GetCf2PdfDetails returned no usable data for claimId={claim_id}."
            )

        # Part III patient/member signature name must come from Beacon's
        # authoritative CF1 record, not from the uploaded automation data.
        cf1_summary = cf2_api.get_cf1_summary(claim_id)
        patient_fullname = str(
            (cf1_summary or {}).get("patientFullname") or ""
        ).strip()
        if not patient_fullname:
            raise cf2_api.Cf2ApiError(
                f"GetPHICCF1Summary returned no patientFullname "
                f"for claimId={claim_id}."
            )

        billing_clerk_name = self._get_billing_clerk_name()
        designation = self._get_official_capacity_designation()

        # NewPdfClaimFormTwo renders the timestamp in Philippine local time.
        # To display the claim's local calendar date (e.g. 07-01-2026), send
        # local midnight as the previous UTC day at 16:00, without milliseconds.
        date_signed_iso = cf2_api.to_utc_midnight_iso(
            data.last_treatment
        ).replace(".000Z", "Z")

        # Part III - member/patient/authorized representative.
        base["sigOverPrintedNameOfAuthRep"] = patient_fullname
        base["patientFullname"] = patient_fullname

        # Part IV - authorized HCI representative.
        base["signatureOverPrintedNameOfAuthHCIRep"] = billing_clerk_name
        base["officialCapacityDesignation"] = designation
        base["dateSigned"] = date_signed_iso

        print("Saving Claim Form 2 PDF fields via API...")
        print(
            "  Part III: "
            f"patientFullname='{patient_fullname}', "
            f"sigOverPrintedNameOfAuthRep='{patient_fullname}'"
        )
        print(
            "  Part IV: "
            f"signature='{billing_clerk_name}', "
            f"designation='{designation}', "
            f"dateSigned='{date_signed_iso}'"
        )
        cf2_api.new_pdf_claim_form_two(claim_id, base)
        print("Claim Form 2 PDF fields saved via API.")
        return True




    def _add_doctor(self, data):
        ids = self._get_ids()
        existing = cf2_api.get_doctors(ids["claim_id"])
        if existing:
            print("Doctor already exists (via API) - skipping creation.")
            return

        client_id = cf2_api.get_client_id()
        hospital_identity = cf2_api.get_hospital_identity(ids["transmittal_id"])
        accreditation_number = data.accreditation_no.strip()
        admission_date_str = data.first_treatment.strftime("%m-%d-%Y")
        discharge_date_iso = cf2_api.to_utc_midnight_iso(data.last_treatment)
        sign_date_str = data.last_treatment.strftime("%m-%d-%Y")

        cf2_api.add_doctor(
            ids["claim_id"],
            client_id,
            accreditation_number,
            sign_date_str,
            admission_date_str,
            discharge_date_iso,
            hospital_identity,
        )
        print("Doctor added (via API).")



    # --------------------------------------------------
    # Save CF2
    # --------------------------------------------------

    def _fill_statement_of_account(self, data):
        """Save the Statement of Account Signatories entirely through API.

        Signatories are populated and saved through the API. We first read the member mobile from CF1, then
        read the existing eSOA signatory record to obtain its ``id``, build
        the same payload Beacon sends from the Signatories SAVE button, POST
        it to Update-signatories, and finally GET the record again to verify
        the values persisted.
        """
        ids = self._get_ids()
        claim_id = ids["claim_id"]


        member_mobile = None

        def _get_member_mobile_from_api():
            nonlocal member_mobile
            cf1_summary = cf2_api.get_cf1_summary(claim_id)
            member_mobile = (cf1_summary or {}).get("memberMobileNumber")
            if not member_mobile:
                raise cf2_api.Cf2ApiError(
                    f"GetPHICCF1Summary returned no memberMobileNumber "
                    f"for claimId={claim_id}."
                )
            member_mobile = str(member_mobile).strip()
            print(f"Member mobile from CF1 API: {member_mobile}")

        self._step(
            "Getting member mobile from CF1 API...",
            _get_member_mobile_from_api,
            critical=True,
        )

        billing_clerk_name = self._get_billing_clerk_name()
        billing_clerk_cp = self._get_billing_clerk_cp()
        date_str = data.last_treatment.strftime("%m-%d-%Y")

        signatories = cf2_api.get_esoa_signatories(claim_id)
        if not signatories:
            raise cf2_api.Cf2ApiError(
                f"GetEsoaSignatories returned no record for phicClaimId={claim_id}."
            )

        signatory_id = signatories.get("id")
        if signatory_id is None:
            raise cf2_api.Cf2ApiError(
                f"GetEsoaSignatories returned no id for phicClaimId={claim_id}."
            )

        payload = {
            "phicClaimId": claim_id,
            "id": signatory_id,
            "preparedBy": billing_clerk_name,
            "adminContactNo": billing_clerk_cp,
            "adminDateSigned": date_str,
            "patientRepresentative": data.patient_name,
            "relationshipOfRepresentative": "",
            "representativeContactNo": member_mobile,
            "representativeDateSigned": date_str,
        }

        print("Updating Statement of Account Signatories via API...")
        cf2_api.update_esoa_signatories(payload)
        print("Statement of Account Signatories saved via API.")

        # Verify the persisted backend record rather than relying on DOM
        # values. This also catches a successful HTTP response that did not
        # actually persist one of the submitted fields.
        saved = cf2_api.get_esoa_signatories(claim_id)
        if not saved:
            raise cf2_api.Cf2ApiError(
                f"GetEsoaSignatories returned no record after save for "
                f"phicClaimId={claim_id}."
            )

        expected = {
            "preparedBy": billing_clerk_name,
            "adminContactNo": billing_clerk_cp,
            "patientRepresentative": data.patient_name,
            "representativeContactNo": member_mobile,
        }
        for field, expected_value in expected.items():
            actual_value = str(saved.get(field) or "").strip()
            if actual_value != str(expected_value or "").strip():
                raise cf2_api.Cf2ApiError(
                    f"Signatories verification failed for {field}: "
                    f"expected '{expected_value}', got '{actual_value}'."
                )

        print("Statement of Account Signatories verified via API.")

        self._step(
            "Saving Claim Form 2 PDF fields via API...",
            lambda: self._fill_claim_form_two_via_api(claim_id, data),
            critical=True,
        )
        print("Claim Form 2 completed via API.")
