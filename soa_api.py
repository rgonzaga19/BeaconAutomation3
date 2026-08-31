"""API layer for Beacon SOA automation.

Business rules stay in soa_automation.py.  This module only mirrors the
HTTP calls observed in Beacon's successful SOA workflow.
"""
from pathlib import Path
from datetime import date, datetime, timedelta
import requests
import browser_session

BASE_URL = "https://beacon-s4.bizbox.ph"
ECLAIMS_API_BASE = "https://eclaimsapi-s4.azurewebsites.net/api/EClaims/v3"

class SoaApiError(RuntimeError): pass

def _headers(json=True):
    token = browser_session.get_auth_token()
    if not token: raise SoaApiError("Beacon auth token is unavailable")
    h={"Authorization":f"Bearer {token}"}
    if json: h["Content-Type"]="application/json"
    return h

def _check(r):
    try: r.raise_for_status()
    except requests.HTTPError as e:
        raise SoaApiError(f"{r.request.method} {r.url} failed ({r.status_code}): {r.text[:1000]}") from e
    return r

def _json(r):
    _check(r)
    if not r.text: return None
    try: return r.json()
    except ValueError: return r.text.strip('"')

def get_transmittal(transmittal_no, client_id=263):
    # Match Beacon's Transmittals search request from the captured UI flow.
    # The endpoint expects the table's date window and paging parameters even
    # when an exact transmittal number is supplied in ``que``.
    today = datetime.now().date()
    date_from = today - timedelta(days=31)
    date_to = today + timedelta(days=1)

    r=requests.get(
        BASE_URL+"/api/PHICTransmittal/GetAllPHICTransmittal",
        headers=_headers(),
        params={
            "clientId": client_id,
            "dateFrom": date_from.strftime("%Y-%m-%dT16:00:00.000Z"),
            "dateTo": date_to.strftime("%Y-%m-%dT15:59:59.999Z"),
            "itemStart": 0,
            "itemEnd": 30,
            "que": str(transmittal_no),
            "transmittalPackageType": 7,
        },
        timeout=30,
    )
    data=_json(r) or {}; rows=data.get("transmittalList") or []
    exact=[x for x in rows if str(x.get("transmittalNumber",""))==str(transmittal_no)]
    return exact[0] if exact else None

def get_claims(transmittal_id):
    return _json(requests.get(BASE_URL+"/api/PHICClaim/GetAllPHICClaimByPHICTransmittalId",headers=_headers(),params={"transmittalId":transmittal_id},timeout=30))

def get_claim(claim_id):
    return _json(requests.get(BASE_URL+"/api/PHICClaim/GetPHICClaim",headers=_headers(),params={"id":claim_id},timeout=30))

def get_cf1(claim_id):
    return _json(requests.get(BASE_URL+"/api/PHICCF1/GetPHICCF1Summary",headers=_headers(),params={"id":claim_id},timeout=30))

def get_charges(claim_id):
    """Return Beacon's current MED and XLSO charge rows exactly as exposed by the claim."""
    meds = _json(requests.get(
        BASE_URL + "/api/PHICChargesDrugAndMedicineController/GetPHICChargesDrugsAndMedicines",
        headers=_headers(),
        params={"phicClaimId": claim_id},
        timeout=30,
    )) or []
    xlso = _json(requests.get(
        BASE_URL + "/api/PHICChargesXLSOController/GetPHICChargesXLSO",
        headers=_headers(),
        params={"phicClaimId": claim_id},
        timeout=30,
    )) or []

    # Normalize the only response shapes we have reason to accept here.
    # The captured successful HAR returns a bare JSON list.
    if isinstance(meds, dict):
        meds = (
            meds.get("items")
            or meds.get("data")
            or meds.get("result")
            or meds.get("charges")
            or []
        )
    if isinstance(xlso, dict):
        xlso = (
            xlso.get("items")
            or xlso.get("data")
            or xlso.get("result")
            or xlso.get("charges")
            or []
        )

    return meds if isinstance(meds, list) else [], xlso if isinstance(xlso, list) else []


def get_documents(claim_id):
    """Return claim documents; ESA document existence is separate from charge-import state."""
    docs = _json(requests.get(
        BASE_URL + "/api/PHICDocument/GetPHICDocuments",
        headers=_headers(),
        params={"phicClaimId": claim_id},
        timeout=30,
    )) or []
    if isinstance(docs, dict):
        docs = docs.get("items") or docs.get("data") or docs.get("documents") or []
    return docs if isinstance(docs, list) else []


def get_soa_state(claim_id):
    """Detect SOA import state and generated ESA-document state independently.

    Primary import signal: existing MED rows (same rule as the original UI automation).
    Recovery signal: if MED rows are unavailable but GetSummary still contains a positive
    actualCharges total, Beacon still retains the imported charge state.

    ESA document presence is intentionally independent and comes only from
    GetPHICDocuments(documentType == "ESA"). A stale ESA document by itself does NOT
    mean the workbook charges are still imported.
    """
    meds, xlso = get_charges(claim_id)

    docs = get_documents(claim_id)
    esa_docs = [
        d for d in docs
        if str(d.get("documentType") or "").strip().upper() == "ESA"
    ]

    imported = bool(meds)
    import_source = "MED endpoint" if imported else None

    summary = None
    summary_actual_total = 0.0

    # IMPORTANT: ESA document existence is NOT proof that charge rows still
    # exist.  Beacon can leave a stale ESA document after the SOA charges have
    # been manually deleted.  In that state MED/XLSO and the SOA summary are
    # empty, and ValidateESOA fails because ItemizedBillingItems is empty.
    #
    # Therefore ESA_document_exists is tracked independently and must never
    # suppress a fresh workbook import.
    if not imported:
        summary = get_summary(claim_id) or {}
        fees = summary.get("feesSummary") or []
        try:
            summary_actual_total = sum(float(r.get("actualCharges") or 0) for r in fees)
        except (TypeError, ValueError):
            summary_actual_total = 0.0

        if summary_actual_total > 0:
            imported = True
            import_source = "SOA summary"

    return {
        "charges_imported": imported,
        "import_source": import_source,
        "med_count": len(meds),
        "xlso_count": len(xlso),
        "summary_actual_total": summary_actual_total,
        "esa_document_exists": bool(esa_docs),
        "esa_document_count": len(esa_docs),
        "summary": summary,
    }


def _upload_file(path, endpoint, claim_id=None):
    path=Path(path); params={"phicClaimId":claim_id} if claim_id is not None else None
    with path.open("rb") as f:
        r=requests.post(BASE_URL+endpoint,headers=_headers(json=False),params=params,
          files={"file_0":(path.name,f,"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},timeout=60)
    return _json(r)

def verify_excel(path): return _upload_file(path,"/api/PHICChargesDrugAndMedicineController/VerifyImportToExcel")
def upload_payment(path, claim_id): return _upload_file(path,"/api/PHICPaymentOfficialReceiptController/UploadPayment",claim_id)
def upload_payment_item(path, claim_id): return _upload_file(path,"/api/PHICPaymentOfficialReceiptController/UploadPaymentItem",claim_id)

def _json_safe(value):
    """Convert only JSON-incompatible workbook date values before requests serializes them."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    return value

def _payload_text(value):
    """Match Beacon's browser payload: blank text fields are sent as empty strings."""
    if value is None:
        return ""
    return str(value).strip()


def _payload_number(value, field_name):
    """Keep Beacon charge numbers numeric and fail clearly if a required value is absent."""
    if value in (None, ""):
        raise SoaApiError(f"SOA workbook is missing required numeric field: {field_name}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise SoaApiError(
            f"SOA workbook has invalid numeric value for {field_name}: {value!r}"
        ) from exc
    return int(number) if number.is_integer() else number


def _payload_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return _payload_text(value)


def _medicine_payload_row(row):
    """Build the exact MED object shape captured from Beacon's successful UI import."""
    return {
        "itemId": _payload_text(row.get("itemId")),
        "brandName": _payload_text(row.get("brandName")),
        "genericName": _payload_text(row.get("genericName")),
        "quantity": _payload_number(row.get("quantity"), "quantity"),
        "price": _payload_number(row.get("price"), "price"),
        "totalCost": _payload_number(row.get("totalCost"), "totalCost"),
        "preparation": _payload_text(row.get("preparation")),
        "renderDate": _payload_date(row.get("renderDate")),
        "instructionFrequency": _payload_text(row.get("instructionFrequency")),
        "route": _payload_text(row.get("route")),
        "drugDescription": _payload_text(row.get("drugDescription")),
        "unitDescription": _payload_text(row.get("unitDescription")),
        "dosage": _payload_text(row.get("dosage")),
        "dosageUnit": _payload_text(row.get("dosageUnit")),
    }


def _xlso_payload_row(row):
    """Build the exact XLSO object shape captured from Beacon's successful UI import."""
    return {
        "type": _payload_text(row.get("type")),
        "description": _payload_text(row.get("description")),
        "quantity": _payload_number(row.get("quantity"), "quantity"),
        "price": _payload_number(row.get("price"), "price"),
        "total": _payload_number(row.get("total"), "total"),
        "renderDate": _payload_date(row.get("renderDate")),
        "esoaGroupType": _payload_text(row.get("esoaGroupType")),
        "unitDescription": _payload_text(row.get("unitDescription")),
        "itemName": _payload_text(row.get("itemName")),
    }


def batch_upload_medicines(claim_id, rows):
    # The workbook can contain footer/separator rows that partially match the
    # detected MED header. Beacon's browser importer ignores those rows.
    valid_rows = [
        row for row in rows
        if row.get("itemId") not in (None, "")
        and row.get("quantity") not in (None, "")
        and row.get("price") not in (None, "")
        and row.get("totalCost") not in (None, "")
    ]
    if not valid_rows:
        raise SoaApiError("SOA workbook contains no complete MED charge rows")

    payload = [_medicine_payload_row(row) for row in valid_rows]
    return _json(requests.post(
        BASE_URL + "/api/PHICChargesDrugAndMedicineController/BatchUploadCharges",
        headers=_headers(),
        params={"phicClaimId": claim_id},
        json=payload,
        timeout=60,
    ))


def batch_upload_xlso(claim_id, rows):
    # Same behavior as Beacon's importer: ignore partial/footer rows and submit
    # only complete XLSO charge records.
    valid_rows = [
        row for row in rows
        if row.get("type") not in (None, "")
        and row.get("description") not in (None, "")
        and row.get("quantity") not in (None, "")
        and row.get("price") not in (None, "")
        and row.get("total") not in (None, "")
    ]
    if not valid_rows:
        raise SoaApiError("SOA workbook contains no complete XLSO charge rows")

    payload = [_xlso_payload_row(row) for row in valid_rows]
    return _json(requests.post(
        BASE_URL + "/api/PHICChargesXLSOController/BatchUploadXLSO",
        headers=_headers(),
        params={"phicClaimId": claim_id},
        json=payload,
        timeout=60,
    ))
def get_summary(claim_id):
    return _json(requests.get(BASE_URL+"/api/PHICEsoa/GetSummary",headers=_headers(),params={"PHICClaimId":claim_id},timeout=30))
def update_summary(payload):
    return _json(requests.post(BASE_URL+"/api/PHICEsoa/UpdateSummary",headers=_headers(),json=payload,timeout=30))
def get_esoa_xml(claim_id, facility_id=263):
    return _json(requests.get(BASE_URL+"/api/PHICEsoa/GetESOAXML",headers=_headers(),params={"claimId":claim_id,"facilityId":facility_id},timeout=30))
def validate_esoa(payload):
    return _json(requests.post(ECLAIMS_API_BASE+"/ValidateESOA",headers=_headers(),json=payload,timeout=60))
def generate_and_upload_esoa(claim_id, facility_id=263):
    return _json(requests.post(BASE_URL+"/api/PHICEsoa/GenerateAndUploadEsoaXML",headers=_headers(),json={"claimId":claim_id,"facilityId":facility_id,"isUpload":True},timeout=60))

def parse_soa_workbook(path):
    """Parse each patient's SOA workbook dynamically from its own row values.

    The workbook template uses human-readable column names, while Beacon's
    batch APIs use different JSON field names. This parser maps the template
    headers to Beacon fields without hardcoding patient-specific values.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    meds = []
    xlso = []

    def norm(value):
        return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())

    med_aliases = {
        "itemid": "itemId",
        "brandname": "brandName",
        "genericname": "genericName",
        "qty": "quantity",
        "quantity": "quantity",
        "price": "price",
        "totalcost": "totalCost",
        "preparation": "preparation",
        "renderdate": "renderDate",
        "frequency": "instructionFrequency",
        "instructionfrequency": "instructionFrequency",
        "route": "route",
        "philhealtmapping": "drugDescription",
        "philhealthmapping": "drugDescription",
        "drugdescription": "drugDescription",
        "philhealthunitmapping": "unitDescription",
        "unitdescription": "unitDescription",
        "dosage": "dosage",
        "dosageunit": "dosageUnit",
    }

    xlso_aliases = {
        "type": "type",
        "description": "description",
        "qty": "quantity",
        "quantity": "quantity",
        "price": "price",
        "total": "total",
        "renderdate": "renderDate",
        "esoagrouptype": "esoaGroupType",
        "philhealthunitmapping": "unitDescription",
        "unitdescription": "unitDescription",
        "philhealthmapping": "itemName",
        "itemname": "itemName",
    }

    def clean_number(value):
        if value in (None, ""):
            return value
        try:
            number = float(value)
            return int(number) if number.is_integer() else number
        except (TypeError, ValueError):
            return value

    def clean_date(value):
        if value in (None, ""):
            return value
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        for header_index, header_row in enumerate(rows[:30]):
            normalized = [norm(cell) for cell in header_row]
            med_score = sum(key in med_aliases for key in normalized)
            xlso_score = sum(key in xlso_aliases for key in normalized)

            if max(med_score, xlso_score) < 4:
                continue

            if med_score > xlso_score:
                aliases = med_aliases
                target = meds
                kind = "MED"
            else:
                aliases = xlso_aliases
                target = xlso
                kind = "XLSO"

            mapping = {
                column_index: aliases[key]
                for column_index, key in enumerate(normalized)
                if key in aliases
            }

            for values in rows[header_index + 1:]:
                if not any(value not in (None, "") for value in values):
                    continue

                obj = {
                    api_field: values[column_index]
                    for column_index, api_field in mapping.items()
                    if column_index < len(values)
                }

                if kind == "MED":
                    if obj.get("itemId") in (None, ""):
                        continue

                    for key in ("quantity", "price", "totalCost"):
                        obj[key] = clean_number(obj.get(key))
                    obj["renderDate"] = clean_date(obj.get("renderDate"))

                    generic = str(obj.get("genericName") or "").strip()
                    obj["genericName"] = generic[:50]
                    obj.setdefault("dosage", "")
                    obj.setdefault("dosageUnit", "")

                else:
                    if obj.get("type") in (None, "") or obj.get("description") in (None, ""):
                        continue

                    obj["type"] = str(obj["type"]).strip().upper()
                    for key in ("quantity", "price", "total"):
                        obj[key] = clean_number(obj.get(key))
                    obj["renderDate"] = clean_date(obj.get("renderDate"))

                target.append(obj)

            break

    if not meds and not xlso:
        raise SoaApiError("SOA workbook could not be mapped to Beacon charge columns")

    return meds, xlso

