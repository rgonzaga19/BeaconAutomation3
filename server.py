"""
Local backend server for the Beabots Electron front end.

This wraps the EXISTING business/automation logic — license.py,
settings.py's load_settings/save_settings, cf2_automation.py, cf2_mapper.py,
soa_automation.py, patient_record.py, date_parser.py, logger.py — behind a
small HTTP + WebSocket API. No automation/business logic was changed here;
only *how results reach the UI* changed: structured JSON responses and
socket events instead of directly writing into tkinter widgets.

Native OS things (file-open dialogs, folder pickers, save-as dialogs) are
NOT handled here — those stay in Electron's main process, which already
has better native dialog support than a browser does. This server only
ever receives paths that Electron has already resolved.

Run directly for local testing:
    python server.py
Electron's main process spawns this exact script (or its PyInstaller-built
exe) as a child process on app launch.
"""

import os
import sys
import threading
from pathlib import Path

from flask import Flask, request, jsonify, send_file
from flask_socketio import SocketIO
from flask_cors import CORS

from openpyxl import load_workbook

from license import validate_license, LicenseError
from settings import load_settings, save_settings
from logger import logger
from patient_record import PatientRecord
from date_parser import parse_dates
from cf2_mapper import build_cf2_data
from cf2_automation import CF2Automation
from soa_automation import SOAAutomation
from beacon import run as beacon_run
from reports import report

app = Flask(__name__)
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*")

# Mirrors the old module-level globals (selected_file / patient_records)
# that used to live in cf2_window.py.
_state = {
    "selected_file": None,
    "patient_records": [],
    # "new_draft" (default) or "existing_draft" — set by the most recent
    # /api/cf2/upload call, read again by /api/cf2/start so the CF2
    # automation knows whether to create a fresh draft or search for one
    # that already exists. See cf2_automation.py's CF2Automation.mode.
    "cf2_mode": "new_draft",
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Defaults mirror exactly what beacon.py's `auto_encode_cf4` branch used to
# hardcode inline, now expanded to cover the full "Pertinent Signs and
# Symptoms" and "Physical Examination" checklists from Beacon's live CF4
# form. Keep this in lockstep with DEFAULT_CF4_SETTINGS in js/cf4.js —
# same keys, same values — since that file renders the form these keys
# back onto and both need to agree with beacon.py's DEFAULT_CF4_DATA on
# what each key means.
# Defaults mirror exactly what beacon.py's `auto_encode_cf4` branch used to
# hardcode inline, now expanded to cover the full "Pertinent Signs and
# Symptoms" and "Physical Examination" checklists from Beacon's live CF4
# form. For those two sections, each key IS the confirmed Beacon `name`
# attribute (camelCase) — beacon.py reads cf4_data[key] and locates
# input[name="key"] directly, no separate translation table. Keep this in
# lockstep with DEFAULT_CF4_SETTINGS in js/cf4.js — same keys, same
# values — since that file renders the form these keys back onto and both
# need to agree with beacon.py's DEFAULT_CF4_DATA on what each key means.
DEFAULT_CF4_SETTINGS = {
    "chief_complaint": "FOR HEMODIALYSIS",
    "history_of_present_illness": "N/A",
    "pertinent_past_medical_history": "N/A",
    "general_survey_awake_alert": True,
    "course_in_ward_order": "UF GOAL MET AT L",

    # Pertinent Signs and Symptoms
    "alteredMentalSensorium": False,
    "abdominalCrampPain": False,
    "anorexia": False,
    "bleedingGums": False,
    "bodyWeakness": True,
    "blurringOfVision": False,
    "chestPainDiscomfort": False,
    "constipation": False,
    "cough": False,
    "diarrhea": False,
    "dizziness": False,
    "dysphagia": False,
    "dyspnea": False,
    "dysuria": False,
    "epistaxis": False,
    "fever": False,
    "frequencyOfUrination": False,
    "headache": False,
    "hematemesis": False,
    "hematuria": False,
    "hemoptysis": False,
    "irritability": False,
    "jaundice": False,
    "lowerExtremityEdema": True,
    "myalgia": False,
    "orthopnea": False,
    "pain": False,
    "painSpecify": "",
    "palpitations": False,
    "seizure": False,
    "skinRashes": False,
    "stoolBloodyBlackTarryMucoid": False,
    "sweating": False,
    "urgency": False,
    "vomiting": False,
    "weightLoss": False,
    "others": False,
    "othersSpecify": "",

    # Physical Examination — HEENT
    "heEssentiallyNormal": True,
    "heSunkenFontanelle": False,
    "heAbnormalPupillaryReaction": False,
    "heOthersChk": False,
    "heOthers": "",
    "heCervicalLympadenopathy": False,
    "heDryMucousMembrane": False,
    "heIctericSclerae": False,
    "hePaleConjunctivae": False,
    "heSunkenEyeballs": False,

    # Physical Examination — Chest / Lungs
    "clEssentiallyNormal": True,
    "clOthersChk": False,
    "clOthers": "",
    "clAsymmetricalChestExpansion": False,
    "clDecreasedBreathSounds": False,
    "clWheezes": False,
    "clLumpsOverBreast": False,
    "clCracklesRales": False,
    "clRetractions": False,

    # Physical Examination — CVS
    "cvEssentiallyNormal": True,
    "cvOthersChk": False,
    "cvOthers": "",
    "cvDisplacedApexBeat": False,
    "cvHeavesThrills": False,
    "cvPericardialBulge": False,
    "cvIrregularRhythm": False,
    "cvMuffledHeartSounds": False,
    "cvMurmur": False,

    # Physical Examination — Abdomen
    "abEssentiallyNormal": True,
    "abOthersChk": False,
    "abOthers": "",
    "abAbdominalRigidity": False,
    "abAbdominalTenderness": False,
    "abHyperactiveBowelSounds": False,
    "abPalpableMasses": False,
    "abTympaniticDullAbdomen": False,
    "abUterineContraction": False,

    # Physical Examination — GU (IE)
    "guEssentiallyNormal": False,
    "guBloodStainedInExamFinger": False,
    "guCervicalDilatation": False,
    "guPresenceofAbnormalDischarge": False,
    "guOthersChk": True,
    "guOthers": "NOT EXAMINE",

    # Physical Examination — Skin/Extremities
    "seEssentiallyNormal": True,
    "sePoorSkinTurgor": False,
    "seClubbing": False,
    "seRashesPetechiae": False,
    "seColdClammy": False,
    "seWeakPulse": False,
    "seCyanosisMottledSkin": False,
    "seOthersChk": False,
    "seOthers": "",
    "seEdemaSwelling": False,
    "seDecreasedMobility": False,
    "sePaleNailbeds": False,

    # Physical Examination — Neuro-exam
    "neEssentiallyNormal": True,
    "nePoorCoordination": False,
    "neAbnormalGait": False,
    "neOthersChk": False,
    "neOthers": "",
    "neAbnormalPositionSense": False,
    "neAbnormalSensation": False,
    "neAbnormalReflexes": False,
    "nePoorAlteredMemory": False,
    "nePoorMuscleToneStrength": False,
}


def _load_cf4_settings():
    """Current CF4 defaults, merged over DEFAULT_CF4_SETTINGS so a
    settings.json saved before some new field existed doesn't come back
    with that field missing."""
    settings = load_settings()
    return {**DEFAULT_CF4_SETTINGS, **settings.get("cf4", {})}


# ---------------------------------------------------------------------------
# Logging bridge — every automation run still calls logger.<level>(...)
# exactly as before (inside cf2_automation.py / soa_automation.py). We just
# point the callback at a socket emit instead of a tkinter Text widget.
# ---------------------------------------------------------------------------
def _emit_log(message, level=None):
    socketio.emit("log", {"message": message, "level": level or "INFO"})


logger.set_callback(_emit_log)


def resource_path(relative_path):
    """Same PyInstaller-aware path resolution as cf2_window.py's version."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# ---------------------------------------------------------------------------
# License & Settings
# (open_settings()'s tkinter dialog is gone — the renderer calls these
# instead. The show/hide-password and view-access-key toggles are pure UI
# state and don't need a backend call at all.)
# ---------------------------------------------------------------------------
@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify(load_settings())


@app.route("/api/settings", methods=["POST"])
def post_settings():
    """
    Body: {"username": ..., "password": ..., "access_key": ...}
    save_settings() itself still owns the "clear session.json if credentials
    changed" rule — untouched, unchanged from settings.py.
    """
    data = request.get_json(force=True)
    settings = load_settings()
    settings.update(data)
    save_settings(settings)
    return jsonify(settings)


@app.route("/api/cf4/settings", methods=["GET"])
def get_cf4_settings():
    """Backs the CF4 screen's initial load — same values beacon.py's
    Auto Encode CF4 step will use on the next run."""
    return jsonify(_load_cf4_settings())


@app.route("/api/cf4/settings", methods=["POST"])
def post_cf4_settings():
    """
    Body: any subset of DEFAULT_CF4_SETTINGS' keys. Merged over the
    current saved values (not replaced wholesale), then written into
    settings.json under "cf4" via the existing save_settings().
    """
    data = request.get_json(force=True)
    settings = load_settings()
    cf4 = {**DEFAULT_CF4_SETTINGS, **settings.get("cf4", {}), **data}
    settings["cf4"] = cf4
    save_settings(settings)
    return jsonify(cf4)


@app.route("/api/license/validate", methods=["POST"])
def license_validate():
    """
    Re-runs the exact license check every window (CF2, Upload SOA) used to
    run for itself before opening. The renderer calls this once before
    showing the CF2 or Upload SOA screen, same as before.
    """
    try:
        settings = load_settings()
        entered_key = settings.get("access_key", "").strip()
        license_info = validate_license(entered_key)

        if not license_info["valid"]:
            return jsonify({"valid": False, "error": "Invalid or expired license."})

        settings["license_owner"] = license_info["owner"]
        settings["license_plan"] = license_info["plan"]
        settings["license_expiry"] = license_info["expires"]
        save_settings(settings)

        return jsonify({"valid": True, **license_info})

    except LicenseError as e:
        return jsonify({"valid": False, "error": str(e)})
    except Exception as e:
        return jsonify({"valid": False, "error": f"Unable to verify the license.\n\n{e}"})


# ---------------------------------------------------------------------------
# CF2 workbook analysis
# Same computation as cf2_window.py's analyze_workbook() (parse_dates,
# build_cf2_data, member-pin handling) — only the *output* changed, from
# Text-widget inserts to a structured JSON payload.
# ---------------------------------------------------------------------------
def _analyze_workbook(workbook, claim_year, claim_month=None, mode="new_draft"):
    sheet = workbook["Sheet1"]
    records = []

    for row in range(2, sheet.max_row + 1):
        # Column layout is the same in both templates — only what column A
        # MEANS changes: Member PIN for a fresh draft (new_draft, the
        # original CF2_Template.xlsx), or an existing Transmittal No. to
        # search for instead (existing_draft, CF2_Template_ExistingDraft.xlsx).
        # B=NAME, C=DOCTOR, D=ACCREDITATION NO., E=TREATMENT DATES either way.
        patient = sheet[f"B{row}"].value
        if patient is None:
            continue

        identifier = sheet[f"A{row}"].value
        doctor = sheet[f"C{row}"].value
        accreditation = sheet[f"D{row}"].value
        treatment_dates = sheet[f"E{row}"].value

        identifier_str = str(identifier).strip() if identifier is not None else ""

        record = PatientRecord(
            transmittal=identifier_str if mode == "existing_draft" else "",
            patient_name=str(patient),
            doctor=str(doctor),
            accreditation_no=str(accreditation),
            treatment_dates_raw=str(treatment_dates),
            member_pin=identifier_str if mode == "new_draft" else "",
        )

        record.treatment_dates = parse_dates(
            record.treatment_dates_raw, claim_year, claim_month
        )

        if record.treatment_dates:
            record.first_treatment = record.treatment_dates[0]
            record.last_treatment = record.treatment_dates[-1]
            record.total_sessions = len(record.treatment_dates)

        records.append(record)

    return records


def _record_to_dict(record, cf2, mode="new_draft"):
    """Same fields the old log_box printed per patient, as JSON instead of
    text. "identifier"/"identifier_label" replace the old hardcoded
    "member_pin" display line so the CF2 window can show whichever column
    A actually meant for this upload (Member PIN vs. Transmittal No.)."""
    return {
        "identifier_label": "Transmittal No." if mode == "existing_draft" else "Member PIN",
        "identifier": record.transmittal if mode == "existing_draft" else record.member_pin,
        "patient_name": record.patient_name,
        "doctor": record.doctor,
        "accreditation_no": record.accreditation_no,
        "treatment_dates_raw": record.treatment_dates_raw,
        "parsed_dates": [d.strftime("%m-%d-%Y") for d in record.treatment_dates],
        "first_treatment": record.first_treatment.strftime("%m-%d-%Y") if record.first_treatment else None,
        "last_treatment": record.last_treatment.strftime("%m-%d-%Y") if record.last_treatment else None,
        "total_sessions": record.total_sessions,
        "cf2": {
            "transmittal": cf2.transmittal,
            "patient_name": cf2.patient_name,
            "doctor": cf2.doctor,
            "accreditation_no": cf2.accreditation_no,
            "first_treatment": cf2.first_treatment.strftime("%m-%d-%Y") if cf2.first_treatment else None,
            "last_treatment": cf2.last_treatment.strftime("%m-%d-%Y") if cf2.last_treatment else None,
            "total_sessions": cf2.total_sessions,
        },
    }


@app.route("/api/cf2/upload", methods=["POST"])
def cf2_upload():
    """
    Body: {"path": "<file path>", "claim_year": 2026, "claim_month": "June",
           "mode": "new_draft" | "existing_draft"}
    Electron's native file-open dialog already resolved the path — this
    endpoint never receives raw file bytes, just the path on disk.
    """
    data = request.get_json(force=True)
    filename = data.get("path")
    claim_year = int(data.get("claim_year"))
    claim_month_name = data.get("claim_month")
    claim_month = MONTH_NAMES.index(claim_month_name) + 1 if claim_month_name else None
    mode = data.get("mode") if data.get("mode") in ("new_draft", "existing_draft") else "new_draft"

    if not filename or not Path(filename).is_file():
        return jsonify({"error": "File not found."}), 400

    try:
        workbook = load_workbook(filename, data_only=True)
    except Exception as ex:
        return jsonify({"error": str(ex)}), 400

    records = _analyze_workbook(workbook, claim_year, claim_month, mode)
    _state["selected_file"] = filename
    _state["patient_records"] = records
    _state["cf2_mode"] = mode

    payload_records = [_record_to_dict(r, build_cf2_data(r), mode) for r in records]

    return jsonify({
        "sheets": workbook.sheetnames,
        "patient_count": len(records),
        "records": payload_records,
    })


@app.route("/api/cf2/download-template", methods=["GET"])
def cf2_download_template():
    """
    Returns the template bytes for the requested mode (?mode=new_draft,
    the default, or ?mode=existing_draft). Electron's renderer triggers
    this, then the main process's native save-as dialog decides where to
    write it — same end result as the old shutil.copyfile() +
    asksaveasfilename() flow.
    """
    mode = request.args.get("mode") if request.args.get("mode") in ("new_draft", "existing_draft") else "new_draft"

    if mode == "existing_draft":
        template_path = resource_path(os.path.join("templates", "CF2_Template_ExistingDraft.xlsx"))
        download_name = "CF2_Template_ExistingDraft.xlsx"
    else:
        template_path = resource_path(os.path.join("templates", "CF2_Template.xlsx"))
        download_name = "CF2_Template.xlsx"

    return send_file(template_path, as_attachment=True, download_name=download_name)


# ---------------------------------------------------------------------------
# CF2 automation run — identical worker logic to the old
# _run_automation_worker(), just triggered by an HTTP call instead of a
# button's command=, and reporting back over the "cf2_done" socket event
# instead of log_box.after(...).
# ---------------------------------------------------------------------------
_cf2_running = False


def _run_cf2_automation():
    global _cf2_running
    automation = None
    try:
        automation = CF2Automation(
            uploaded_excel_path=_state["selected_file"],
            mode=_state["cf2_mode"],
        )
        for record in _state["patient_records"]:
            try:
                automation.process_patient(record)
            except Exception as ex:
                automation.results.append({
                    "transmittal": getattr(record, "transmittal", "?"),
                    "patient_name": getattr(record, "patient_name", "?"),
                    "status": "failed",
                    "message": f"Unhandled error: {ex}",
                })
    except Exception as ex:
        socketio.emit("log", {"message": f"ERROR: {ex}", "level": "ERROR"})
    finally:
        if automation is not None:
            try:
                automation.close()
            except Exception as ex:
                socketio.emit("log", {
                    "message": f"WARNING: Could not close automation session: {ex}",
                    "level": "WARNING",
                })

        results = automation.get_summary() if automation is not None else []
        _cf2_running = False
        socketio.emit("cf2_done", {"results": results})


@app.route("/api/cf2/start", methods=["POST"])
def cf2_start():
    global _cf2_running
    if not _state["patient_records"]:
        return jsonify({"error": "No patients loaded."}), 400
    if _cf2_running:
        return jsonify({"error": "Automation already running."}), 409

    _cf2_running = True
    threading.Thread(target=_run_cf2_automation, daemon=True).start()
    return jsonify({"started": True})


# ---------------------------------------------------------------------------
# SOA upload automation — identical to upload_soa_window.py's
# start_soa_automation()/run_automation() worker, minus the tkinter widget
# disabling (the renderer handles disabling its own controls).
# ---------------------------------------------------------------------------
_soa_running = False


def _run_soa_automation(soa_folder, transmittals):
    global _soa_running
    try:
        soa_automation = SOAAutomation(soa_folder=soa_folder)
        soa_automation.run(transmittals)
    except Exception as ex:
        socketio.emit("log", {"message": f"FATAL ERROR: {ex}", "level": "ERROR"})
    finally:
        _soa_running = False
        socketio.emit("soa_done", {})


@app.route("/api/soa/start", methods=["POST"])
def soa_start():
    global _soa_running
    data = request.get_json(force=True)
    transmittals = data.get("transmittals", [])
    soa_folder = data.get("soa_folder", "").strip()

    if not transmittals:
        return jsonify({"error": "Please enter at least one transmittal number."}), 400
    if not soa_folder:
        return jsonify({"error": "Please select the folder where your SOA files are located."}), 400
    if not Path(soa_folder).is_dir():
        return jsonify({"error": f"The selected SOA folder does not exist:\n\n{soa_folder}"}), 400
    if _soa_running:
        return jsonify({"error": "Automation already running."}), 409

    # Remember the folder choice for next time — same as the old
    # browse_soa_folder()'s settings["soa_folder"] = chosen; save_settings(settings)
    settings = load_settings()
    settings["soa_folder"] = soa_folder
    save_settings(settings)

    _soa_running = True
    threading.Thread(target=_run_soa_automation, args=(soa_folder, transmittals), daemon=True).start()
    return jsonify({"started": True})


# ---------------------------------------------------------------------------
# Main dashboard automation (beacon.run() + reports.report) — the flow
# ui.py's start_automation()/run_automation() drove. Distinct from the CF2
# and SOA automations above: this is the original "Move to CF2" precursor
# run against a plain list of transmittals.
# ---------------------------------------------------------------------------
_beacon_running = False


def _run_beacon_automation(transmittals, auto_encode_cf4, cf4_settings):
    global _beacon_running
    try:
        beacon_run(
            transmittals,
            auto_encode_cf4=auto_encode_cf4,
            cf4_data=cf4_settings,
        )
    except Exception as ex:
        socketio.emit("log", {"message": f"ERROR: {ex}", "level": "ERROR"})
    finally:
        _beacon_running = False
        socketio.emit("beacon_done", {"results": report.results})


@app.route("/api/beacon/start", methods=["POST"])
def beacon_start():
    global _beacon_running
    data = request.get_json(force=True)
    transmittals = data.get("transmittals", [])
    auto_encode_cf4 = bool(data.get("auto_encode_cf4", False))

    if not transmittals:
        return jsonify({"error": "Please paste at least one transmittal number."}), 400
    if _beacon_running:
        return jsonify({"error": "Automation already running."}), 409

    # Read whatever was last saved on the CF4 screen (falls back to
    # DEFAULT_CF4_SETTINGS for anything never saved) — this is what makes
    # the checkbox on the dashboard use the *current* CF4 defaults rather
    # than whatever was hardcoded in beacon.py at build time.
    cf4_settings = _load_cf4_settings()

    _beacon_running = True
    threading.Thread(
        target=_run_beacon_automation,
        args=(transmittals, auto_encode_cf4, cf4_settings),
        daemon=True,
    ).start()
    return jsonify({"started": True})


if __name__ == "__main__":
    port = int(os.environ.get("BEABOTS_PORT", 5417))
    socketio.run(app, host="127.0.0.1", port=port, allow_unsafe_werkzeug=True)